#Comando que carrega uma pasta de trabalho ja exitente
from openpyxl import Workbook, load_workbook
#Comando que carrega um arquivo xlsx ja exitente a partir da pasta do seu arquivo python
wb = load_workbook('planilha/Grades.xlsx')
#Comando que carrega a planilha ativa da pasta de trabalho
ws = wb.active
#Comando que mostra o valor de determinada celula
print(ws['A2'].value)
#Comando que altera o valor da celular selecionada (não é possivel editar o valor de uma celula com a planilha aperta no excel)
ws['A2'].value = 'Gabriel'
print(ws['A2'].value)
#Comando que cria uma planilha na pasta de trabalho atual
wb.create_sheet('test')
#Comando que salva a planilha
wb.save('planilha/Grades.xlsx')
#Comando que exibe a quantidade de tabelas e o nome delas 
print(wb.sheetnames)