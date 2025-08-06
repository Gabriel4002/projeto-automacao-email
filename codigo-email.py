from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from codigo_envio_email import enviar_email


data = {
    "Gabriel":{
        "Email": "gabriellobato315@gmail.com",
        "Notas": {
        "Matematica": 10,
        "Portugues": 8,
        "Historia": 6,
        "Geografia": 1
        }
    },
    "Ana":{
        "Email": "ana@email.com",
        "Notas": {
        "Matematica": 4,
        "Portugues": 3,
        "Historia": 7,
        "Geografia": 3
        }
    },
    "Jorge Willem":{
        "Email": "jorgewillem@email.com",
        "Notas": {
        "Matematica": 9,
        "Portugues": 4,
        "Historia": 2,
        "Geografia": 10
        }
    }
}


wb = Workbook()
ws = wb.active
ws.title = "Notas dos Alunos"

headings = ["Nome"] + list(data["Gabriel"]["Notas"].keys())
ws.append(headings)

for aluno, dados in data.items():
    notas = dados["Notas"]
    linha_planilha = [aluno] + list(notas.values())
    ws.append(linha_planilha)

    corpo = f"Olá {aluno},\n\nAqui estão suas notas:\n"
    total = 0
    for materia, nota in notas.items():
        corpo += f"{materia}: {nota}\n"
        total += nota
    media = total / len(notas)
    corpo += f"\nMédia: {media:.2f}\n\nAtenciosamente,\nSecretaria Escolar"

    enviar_email(dados["Email"], "Boletim Escolar", corpo)
    print(f"E-mail enviado para {aluno}")

wb.save("planilha/notas_alunos.xlsx")