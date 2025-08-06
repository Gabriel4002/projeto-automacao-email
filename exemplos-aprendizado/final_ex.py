from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
#Este comando importa 'Font', que é capaz de alterar a formatação das fonte com deixa-la em italico, negrito, mudar sua cor
from openpyxl.styles import Font

data = {
    "Gabriel":{
        "Email": "gabriellobato410@gmail.com",
        "Notas": {
        "Matematica": 10,
        "Portugues": 8,
        "Historia": 6,
        "Geografia": 1
        }
    },
    "Ana":{
        "Email": "ana@email.com",
        "Notas": {
        "Matematica": 4,
        "Portugues": 3,
        "Historia": 7,
        "Geografia": 3
        }
    },
    "Jorge Willem":{
        "Email": "jorgewillem@email.com",
        "Notas": {
        "Matematica": 9,
        "Portugues": 4,
        "Historia": 2,
        "Geografia": 10
        }
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Notas"

#Adiciona o cabecalho das colunas
headings = ["Nome"] + list(data["Gabriel"].keys())
ws.append(headings)

#Um loop que me fornece as notas das pessoas em determindas materias
for person in data:
    nota = list(data[person].values())
    #Ira informar o nome da pessoa e as notas dela
    ws.append([person] + nota )

for col in range(2, len(data["Gabriel"]) + 2):
    char = get_column_letter(col)
    ws[char + "5"] = f"=SUM({char + '2'}:{char + '4'})/{len(data["Notas"])}"

#Loop que transforma a primeira linha das colunas especificadas em negrito
for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True)
wb.save("planilha/exfinal.xlsx")