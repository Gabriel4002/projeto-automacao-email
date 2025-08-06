from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('planilha/teste.xlsx')
ws = wb.active

#Esta função mescla as celulas do intervalo estipulado
ws.merge_cells('A1:G1')

#Esta função desmescla as celulas
ws.unmerge_cells('A1:G1')

#Esta função insere uma nova linha abaixo da linha especificada
ws.insert_rows(3)

#Esta função deleta a linha especificada
ws.delete_rows(3)

#Esta função insere colunas (os numeros representam as letras ex: coluna a = 1)
ws.insert_cols(2)

#Esta função deleta colunas
ws.delete_cols(2)

#Esta função muda a localização das celulas selecionadas de acordo com a quantidade de linhas e colunas que você deseja mover, no caso 'rows = 2' move as celulas duas linhas para baixo(se fosse -2 seria duas linhas para cima) e cols = 3 mantem a mesma ideia porem com colunas e se movendo para os lados
ws.move_range("A1:G1", rows=3, cols=2)
wb.save('planilha/teste.xlsx')