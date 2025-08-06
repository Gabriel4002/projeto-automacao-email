from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
#Comando que cria uma pasta de trabalho
wb = load_workbook('planilha/Grades.xlsx')
ws = wb.active
#Umm loop de 1 ate 11 (representando as linhas)
for row in range(1,5):
    #Um loop de 1 até 5 (representando as colunas)
    for col in range(1, 4):
        #Essa função pega um inteiro entre 1 e 26 (letras do alfabeto), e fornece o caractere que ele representa
        char = get_column_letter(col)
        #Essa função unira o valor da coluna (ex: A) e o numero da linha (ex: A1) e exibira o valor dessas colunas
        print(ws[char + str(row)].value )


